from openpyxl import *


def get_names(fileName): #format: [[name, [want1, want2, want3], [dontwant1, dontwant2, dontwant3]]]
    #wb = load_workbook('../Google Group Picker/Isabel Work/Test no preferences.xlsx')
    wb = load_workbook(fileName)
    ws = wb.active
    currentName = ""
    currentRow = 2
    rtn = []
    while currentName != "NoneNone":  #NoneNone = 2 blank cells
        currentRow += 1
        currentName = str(ws['A' + str(currentRow)].value)
        currentName += str(ws['B' + str(currentRow)].value)
        rtn += [[str(ws['A' + str(currentRow)].value) + " " + str(ws['B' + str(currentRow)].value),
               [str(ws['E' + str(currentRow)].value), str(ws['F' + str(currentRow)].value), str(ws['G' + str(currentRow)].value)],
               [str(ws['H' + str(currentRow)].value), str(ws['I' + str(currentRow)].value), str(ws['J' + str(currentRow)].value)]]]
    return rtn
